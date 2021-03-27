import openpyxl
import json
import tkinter as tk
from tkinter.filedialog import askopenfile

# VARIABLES
allData = []
header = 0

root = tk.Tk()
root.withdraw()

file_path = askopenfile(mode='r')


# READ FILE ImportExcel.xlsx
try:
    wb = openpyxl.load_workbook(filename=file_path.name)

    sheet = wb.active

    # Discover how much column has
    for x in range(1,1000):
        value = sheet.cell(row=1, column= x).value
        if value == None:
            break
        header = x


    for x in range(2, 1000):    
        register = []
        for y in range(1, header+1):
            value = sheet.cell(row=x, column=y).value
            if value == None: 
                break
            register.append(value)
        # print(register)
        allData.append(register)
        value = sheet.cell(row=x, column=1).value
        if value == None:
            break


    # CREATE A FORMAT JSON
    x = '['
    for key in allData:
        x = x + '{'
        for c in range(1, header + 1):
            valueCell = sheet.cell(row=1, column=c).value
            try:
                x = x + '"' + str(valueCell) + '":"' + key[c-1] + '",'
            except:
                x = x[:-2]
                break
        x = x[:-1]
        x = x + '},'

    x = x[:-1]
    x = x + ']'

    with open('data.json', 'w') as data:
        json.dump(x, data)

    print("JSON Created...")
except:
    print("Fail to read")

